import streamlit as st
import camelot.io as camelot
import pandas as pd
from io import BytesIO
import tempfile
import os
import re
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


st.set_page_config(page_title="PDF vers Excel", page_icon="📄")
st.title("📄 PDF vers Excel")


def clean_text(value):
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def clean_dataframe(df):
    return df.copy().map(clean_text)


def normalize_text(value):
    value = clean_text(value).lower()
    value = re.sub(r"\s+", " ", value)
    value = re.sub(r"[^a-z0-9à-ÿ()_,.'/%\-]+", "", value)
    return value.strip()


def is_header(row):
    filled = [clean_text(x) for x in row if clean_text(x)]

    if len(filled) < 2:
        return False

    text_count = sum(bool(re.search(r"[A-Za-zÀ-ÿ]", x)) for x in filled)
    number_count = sum(bool(re.search(r"\d", x)) for x in filled)

    return text_count >= 2 and text_count >= number_count


def make_unique_columns(columns):
    seen = {}
    result = []

    for col in columns:
        col = str(col).strip() or "Colonne"

        if col not in seen:
            seen[col] = 0
            result.append(col)
        else:
            seen[col] += 1
            result.append(f"{col}_{seen[col]}")

    return result


def build_headers_from_rows(header_rows):
    max_cols = max(len(row) for row in header_rows)
    headers = []

    for col in range(max_cols):
        parts = []

        for row in header_rows:
            value = clean_text(row[col]) if col < len(row) else ""

            if value and value not in parts:
                parts.append(value)

        headers.append(" ".join(parts).strip() or f"Colonne_{col + 1}")

    return make_unique_columns(headers)


def extract_tables(pdf_path):
    tables = camelot.read_pdf(
        pdf_path,
        pages="all",
        flavor="lattice",
        strip_text="\n"
    )

    result = []

    for table in tables:
        df = clean_dataframe(table.df)

        if not df.empty:
            result.append({
                "df": df,
                "page": int(table.parsing_report.get("page", 0))
            })

    return result


def split_header_and_data(df):
    rows = df.values.tolist()
    header_rows = []

    for i in range(min(3, len(rows))):
        if is_header(rows[i]):
            header_rows.append(rows[i])
        else:
            break

    if not header_rows:
        return None, df.reset_index(drop=True), []

    headers = build_headers_from_rows(header_rows)
    data = df.iloc[len(header_rows):].reset_index(drop=True)

    return tuple(headers), data, header_rows


def merge_tables(raw_tables):
    groups = {}
    last_header_by_cols = {}

    for item in raw_tables:
        df = item["df"]
        header, data, header_rows = split_header_and_data(df)
        col_count = df.shape[1]

        if header is not None:
            header_key = tuple(normalize_text(x) for x in header)
            display_header = header
            display_header_rows = header_rows
            last_header_by_cols[col_count] = (
                header_key,
                display_header,
                display_header_rows
            )
        else:
            header_key, display_header, display_header_rows = last_header_by_cols.get(
                col_count,
                (
                    tuple(f"col_{i}" for i in range(col_count)),
                    tuple(f"Colonne_{i + 1}" for i in range(col_count)),
                    []
                )
            )

        key = f"{col_count}_" + "|".join(header_key)

        if key not in groups:
            groups[key] = {
                "header_key": header_key,
                "display_header": display_header,
                "header_rows": display_header_rows,
                "data": data,
                "pages": [item["page"]],
                "cols": col_count
            }
        else:
            groups[key]["data"] = pd.concat(
                [groups[key]["data"], data],
                ignore_index=True
            )
            groups[key]["pages"].append(item["page"])

    return list(groups.values())


def remove_repeated_headers(df, header_key):
    if header_key is None:
        return df

    header_list = list(header_key)

    def same_as_header(row):
        row_values = [normalize_text(x) for x in row.tolist()]
        return row_values == header_list

    return df[~df.apply(same_as_header, axis=1)]


def build_final_dataframe(table):
    df = table["data"].copy()
    df = remove_repeated_headers(df, table["header_key"])

    headers = list(table["display_header"])

    if len(headers) == df.shape[1]:
        df.columns = make_unique_columns(headers)
    else:
        df.columns = make_unique_columns(
            [f"Colonne_{i}" for i in range(1, df.shape[1] + 1)]
        )

    return df.reset_index(drop=True)


def style_excel_sheet(ws, header_row_count):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="8EAADB")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

    for row in range(1, header_row_count + 1):
        for cell in ws[row]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 24

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 28


def merge_empty_cells_like_pdf(ws, start_row):
    max_row = ws.max_row
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        start = None

        for row in range(start_row, max_row + 1):
            value = ws.cell(row=row, column=col).value

            if value not in [None, ""]:
                if start is not None and row - start > 1:
                    ws.merge_cells(
                        start_row=start,
                        start_column=col,
                        end_row=row - 1,
                        end_column=col
                    )
                start = row

        if start is not None and max_row - start >= 1:
            if all(ws.cell(row=r, column=col).value in [None, ""] for r in range(start + 1, max_row + 1)):
                ws.merge_cells(
                    start_row=start,
                    start_column=col,
                    end_row=max_row,
                    end_column=col
                )


def merge_header_cells(ws, header_row_count):
    if header_row_count <= 1:
        return

    max_col = ws.max_column

    # Fusion horizontale dans l'en-tête
    for row in range(1, header_row_count + 1):
        start_col = None
        current_value = None

        for col in range(1, max_col + 2):
            value = ws.cell(row=row, column=col).value if col <= max_col else None

            if value not in [None, ""]:
                if start_col is not None and col - start_col > 1:
                    ws.merge_cells(
                        start_row=row,
                        start_column=start_col,
                        end_row=row,
                        end_column=col - 1
                    )

                start_col = col
                current_value = value

        if start_col is not None and max_col - start_col >= 1:
            empty_after = all(
                ws.cell(row=row, column=c).value in [None, ""]
                for c in range(start_col + 1, max_col + 1)
            )

            if empty_after:
                ws.merge_cells(
                    start_row=row,
                    start_column=start_col,
                    end_row=row,
                    end_column=max_col
                )

    # Fusion verticale dans l'en-tête
    for col in range(1, max_col + 1):
        top_value = ws.cell(row=1, column=col).value

        if top_value not in [None, ""]:
            below_empty = all(
                ws.cell(row=r, column=col).value in [None, ""]
                for r in range(2, header_row_count + 1)
            )

            if below_empty:
                ws.merge_cells(
                    start_row=1,
                    start_column=col,
                    end_row=header_row_count,
                    end_column=col
                )


uploaded_file = st.file_uploader("Importer un PDF", type=["pdf"])

if uploaded_file is not None:
    if st.button("Extraire les tableaux"):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(uploaded_file.read())
            pdf_path = tmp.name

        try:
            with st.spinner("Extraction des tableaux..."):
                raw_tables = extract_tables(pdf_path)
                merged_tables = merge_tables(raw_tables)

            if not merged_tables:
                st.error("Aucun tableau détecté.")
            else:
                st.success(f"{len(merged_tables)} tableau(x) distinct(s) trouvé(s).")

                output = BytesIO()

                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    for i, table in enumerate(merged_tables, start=1):
                        final_df = build_final_dataframe(table)
                        sheet_name = f"Tableau_{i}"[:31]

                        header_rows = table["header_rows"]
                        header_row_count = len(header_rows)

                        final_df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            index=False,
                            header=False,
                            startrow=header_row_count
                        )

                        ws = writer.book[sheet_name]

                        for r, row in enumerate(header_rows, start=1):
                            for c, value in enumerate(row, start=1):
                                ws.cell(row=r, column=c).value = value

                        if header_row_count == 0:
                            for c, col_name in enumerate(final_df.columns, start=1):
                                ws.cell(row=1, column=c).value = col_name
                            header_row_count = 1

                        merge_header_cells(ws, header_row_count)
                        merge_empty_cells_like_pdf(ws, header_row_count + 1)
                        style_excel_sheet(ws, header_row_count)

                        st.subheader(f"{sheet_name} - pages {table['pages']}")
                        st.dataframe(final_df.astype(str))

                output.seek(0)

                st.download_button(
                    label="📥 Télécharger Excel avec forme",
                    data=output,
                    file_name="tableaux_avec_forme.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        finally:
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
